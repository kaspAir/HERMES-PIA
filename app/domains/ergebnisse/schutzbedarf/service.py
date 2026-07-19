"""Schutzbedarfsanalyse – befüllt das BACS-Excel-Template aus dem PIA.

Schreibt NUR Eingabezellen (Deckblatt/Informationsverzeichnis deterministisch +
beratende Beurteilung in Tab 4). Alle Formeln (Einstufung) bleiben unangetastet.
"""
import io
import json
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from app.domains.ergebnisse.models import ErgebnisEntwurf
from app.domains.ergebnisse.projektwissen import Projektwissen
from app.domains.ergebnisse.schutzbedarf import cellmap as CM
from app.domains.ergebnisse.schutzbedarf.proposals import (
    auswirkungen,
    deckblatt_und_gruppen,
    erhebung,
)
from app.domains.projekt.reference import ERG_PIA
from app.shared.database import SessionLocal

METHOD_ID = "schutzbedarfsanalyse"


class SchutzbedarfService:
    def __init__(self, interview_service, projekt_service, methods_dir, llm=None):
        self.interview = interview_service
        self.projekte = projekt_service
        self.methods_dir = Path(methods_dir)
        self.llm = llm

    def _template_pfad(self):
        return self.methods_dir / METHOD_ID / "template" / CM.TEMPLATE_DATEI

    # ---- PIA-Zugriff (nur lesen) ---------------------------------------- #
    def _pia(self, projekt):
        for erg in self.projekte.ergebnisse(projekt.id):
            if erg.ergebnistyp == ERG_PIA:
                s = self.interview.session_for_ergebnis(erg.id)
                if s and s.answers_json:
                    return json.loads(s.answers_json), s
        return {}, None

    def _metadata(self, projekt, session):
        pl = (session.created_by if session else None) or ""
        return {
            "projektname": projekt.name or "",
            "projektnummer": projekt.projektnummer or "",
            "verwaltungseinheit": projekt.verwaltungseinheit or "",
            "auftraggeber": projekt.auftraggeber or (session.auftraggeber if session else "") or "",
            "projektleiter": pl,
        }

    def projektwissen(self, projekt):
        pia, session = self._pia(projekt)
        return Projektwissen(pia, metadata=self._metadata(projekt, session)), session

    # ---- Template-Analyse ----------------------------------------------- #
    @staticmethod
    def _erhebung_input_cells(ws):
        """Alle Eingabezellen (mit 'Trifft'-Dropdown) in Tab 4 – doppelte Absicherung."""
        cells = set()
        for dv in ws.data_validations.dataValidation:
            if dv.type == "list" and "Trifft" in str(dv.formula1 or ""):
                for rng in dv.sqref.ranges:
                    for row in range(rng.min_row, rng.max_row + 1):
                        for col in range(rng.min_col, rng.max_col + 1):
                            cells.add(f"{get_column_letter(col)}{row}")
        return cells

    @staticmethod
    def _szenarien(ws, input_cells):
        """[(zeile, text)] der Schaden-Szenarien (Spalte B) je Eingabe-Zeile."""
        zeilen = sorted({int("".join(filter(str.isdigit, c))) for c in input_cells})
        out = []
        for z in zeilen:
            t = ws[f"B{z}"].value
            if isinstance(t, str) and t.strip():
                out.append((z, t.strip()))
        return out

    # ---- Zellwerte bauen (Vorschlag) ------------------------------------ #
    def build_cellvalues(self, wissen):
        wb = openpyxl.load_workbook(self._template_pfad())
        ws_erh = wb[CM.TAB_ERHEBUNG]
        input_cells = self._erhebung_input_cells(ws_erh)
        szenarien = self._szenarien(ws_erh, input_cells)

        cv = {CM.TAB_DECKBLATT: {}, CM.TAB_INFOVERZEICHNIS: {},
              CM.TAB_AUSWIRKUNGEN: {}, CM.TAB_ERHEBUNG: {}}
        md = wissen.metadata
        D = cv[CM.TAB_DECKBLATT]
        if md.get("projektname"):
            D[CM.DECKBLATT["schutzobjektname"]] = md["projektname"]
        # Interne Bezeichnung: Projektnummer, sonst Projektname als Referenz.
        if md.get("projektnummer") or md.get("projektname"):
            D[CM.DECKBLATT["interne_bezeichnung"]] = md.get("projektnummer") or md["projektname"]
        if md.get("verwaltungseinheit"):
            D[CM.DECKBLATT["amt"]] = md["verwaltungseinheit"]
        involvierte = ", ".join(x for x in (md.get("auftraggeber"), md.get("projektleiter")) if x)
        if involvierte:
            D[CM.DECKBLATT["involvierte"]] = involvierte

        # LLM (getrennte, kompakte Aufrufe – zuverlässiger als ein grosser).
        info = deckblatt_und_gruppen(wissen, self.llm)
        for feld in ("beschreibung", "geschaeftsprozesse", "zugriff", "geografisch"):
            if info.get(feld):
                D[CM.DECKBLATT[feld]] = str(info[feld])[:900]

        gruppen = info.get("gruppen", [])[:len(CM.INFO_ZEILEN)]
        ausw = auswirkungen(wissen, [g.get("gruppe", "") for g in gruppen], self.llm)
        I = cv[CM.TAB_INFOVERZEICHNIS]
        A = cv[CM.TAB_AUSWIRKUNGEN]
        for idx, g in enumerate(gruppen):
            row = CM.INFO_ZEILEN.start + idx
            if g.get("gruppe"):
                I[f"{CM.INFO_SPALTEN['gruppe']}{row}"] = str(g["gruppe"])[:250]
            if g.get("personendaten"):
                I[f"{CM.INFO_SPALTEN['personendaten']}{row}"] = str(g["personendaten"])[:250]
            # Dropdowns nur bei exakt gültigem Wert setzen.
            if g.get("klassifizierung") in CM.INFO_KLASS_WERTE:
                I[f"{CM.INFO_SPALTEN['klassifizierung']}{row}"] = g["klassifizierung"]
            if g.get("risiko") in CM.INFO_RISIKO_WERTE:
                I[f"{CM.INFO_SPALTEN['risiko']}{row}"] = g["risiko"]
            # Tab 3: Auswirkungstexte je Grundwert (Freitext, nach Gruppenname gematcht).
            impakt = ausw.get(str(g.get("gruppe", "")).strip().lower(), {})
            for gw, spalte in CM.AUSWIRKUNG_SPALTE.items():
                if impakt.get(gw):
                    A[f"{spalte}{row}"] = str(impakt[gw])[:600]

        # Tab 4: beratende Beurteilung – NUR gültige Eingabezellen setzen.
        E = cv[CM.TAB_ERHEBUNG]
        for row, grundwerte in erhebung(wissen, szenarien, self.llm).items():
            for gw in grundwerte:
                coord = f"{CM.GRUNDWERT_SPALTE[gw]}{row}"
                if coord in input_cells:
                    E[coord] = CM.TRIFFT_ZU
        return cv

    # ---- Persistenz ----------------------------------------------------- #
    def get_entwurf(self, projekt_id):
        return SessionLocal().query(ErgebnisEntwurf).filter(
            ErgebnisEntwurf.projekt_id == int(projekt_id),
            ErgebnisEntwurf.ergebnistyp == METHOD_ID,
        ).first()

    def erzeuge_entwurf(self, projekt):
        wissen, _ = self.projektwissen(projekt)
        cv = self.build_cellvalues(wissen)
        db = SessionLocal()
        row = db.query(ErgebnisEntwurf).filter(
            ErgebnisEntwurf.projekt_id == projekt.id,
            ErgebnisEntwurf.ergebnistyp == METHOD_ID,
        ).first()
        if row is None:
            row = ErgebnisEntwurf(projekt_id=projekt.id, ergebnistyp=METHOD_ID)
            db.add(row)
        row.answers_json = json.dumps(cv, ensure_ascii=False, indent=2)
        db.commit()
        db.refresh(row)
        return row

    # ---- Excel erzeugen (Formeln unangetastet) -------------------------- #
    def generate_xlsx(self, projekt):
        entwurf = self.get_entwurf(projekt.id)
        if entwurf and entwurf.answers_json:
            cv = json.loads(entwurf.answers_json)
        else:
            wissen, _ = self.projektwissen(projekt)
            cv = self.build_cellvalues(wissen)
        wb = openpyxl.load_workbook(self._template_pfad())
        erh_cells = self._erhebung_input_cells(wb[CM.TAB_ERHEBUNG])
        for sheet, zellen in cv.items():
            if sheet not in wb.sheetnames:
                continue
            ws = wb[sheet]
            for coord, wert in zellen.items():
                # Sicherheitsnetz: in Tab 4 nur echte Eingabezellen beschreiben.
                if sheet == CM.TAB_ERHEBUNG and coord not in erh_cells:
                    continue
                cell = ws[coord]
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue  # niemals eine Formel überschreiben
                cell.value = wert
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
