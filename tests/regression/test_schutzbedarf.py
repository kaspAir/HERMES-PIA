"""Beweist: Schutzbedarfsanalyse befüllt das BACS-Excel NUR in Eingabezellen –
die Formeln bleiben unangetastet. Modular, PIA unberührt."""
import io
import json

import openpyxl
import pytest

from app.config import Config, get_config
from app.domains.ergebnisse.projektwissen import Projektwissen
from app.domains.ergebnisse.schutzbedarf import cellmap as CM
from app.domains.ergebnisse.schutzbedarf.service import SchutzbedarfService
from app.factory import create_app


class _FakeLLM:
    def __init__(self, payload):
        self._p = payload

    def complete(self, system, messages, max_tokens=1024):
        return json.dumps(self._p)


def _svc(llm=None):
    return SchutzbedarfService(None, None, get_config().METHODS_DIR, llm=llm)


def _formeln(ws):
    return sum(1 for r in ws.iter_rows() for c in r
               if isinstance(c.value, str) and c.value.startswith("="))


def test_template_vorhanden_und_ladbar():
    wb = openpyxl.load_workbook(_svc()._template_pfad())
    assert CM.TAB_ERHEBUNG in wb.sheetnames
    assert _formeln(wb[CM.TAB_ERHEBUNG]) == 127 and _formeln(wb["6. Einstufung"]) == 35


def test_build_und_fuellen_erhaelt_formeln():
    fake = _FakeLLM({
        "beschreibung": "Nachfolgelösung Juris Fiat der Justizbehörden.",
        "gruppen": [{"gruppe": "Straf- und Personendaten",
                     "personendaten": "besonders schützenswerte Personendaten"}],
        "zeilen": [{"zeile": 6, "grundwerte": ["vertraulichkeit", "integritaet"]}],
    })
    svc = _svc(llm=fake)
    wissen = Projektwissen({"ausgangslage": {"extracted": {"text": "Justizsystem-Ablösung."}}},
                           metadata={"projektname": "BKI Test 2", "verwaltungseinheit": "Justiz",
                                     "auftraggeber": "Monika Musterfrau", "projektleiter": "Helene Digital"})
    cv = svc.build_cellvalues(wissen)
    # Deckblatt aus PIA-Metadaten
    assert cv[CM.TAB_DECKBLATT]["D6"] == "BKI Test 2"
    assert "Monika" in cv[CM.TAB_DECKBLATT]["D13"]
    # Tab 4: Vorschlag nur in gültigen Eingabezellen, gültiger Wert
    assert cv[CM.TAB_ERHEBUNG]["C6"] == CM.TRIFFT_ZU     # Vertraulichkeit, Zeile 6
    assert cv[CM.TAB_ERHEBUNG]["E6"] == CM.TRIFFT_ZU     # Integrität, Zeile 6

    # Zellwerte in eine Kopie schreiben (wie generate_xlsx) und Formeln prüfen
    wb = openpyxl.load_workbook(svc._template_pfad())
    for sheet, zellen in cv.items():
        ws = wb[sheet]
        for coord, wert in zellen.items():
            if isinstance(ws[coord].value, str) and ws[coord].value.startswith("="):
                pytest.fail(f"Formelzelle {sheet}!{coord} würde überschrieben")
            ws[coord].value = wert
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    wb2 = openpyxl.load_workbook(buf)
    assert _formeln(wb2[CM.TAB_ERHEBUNG]) == 127          # alle Formeln erhalten
    assert _formeln(wb2["6. Einstufung"]) == 35
    assert wb2[CM.TAB_DECKBLATT]["D6"].value == "BKI Test 2"
    assert wb2[CM.TAB_ERHEBUNG]["C6"].value == CM.TRIFFT_ZU


def test_ohne_llm_nur_deterministisch():
    svc = _svc(llm=None)
    cv = svc.build_cellvalues(Projektwissen({}, metadata={"projektname": "P"}))
    assert cv[CM.TAB_DECKBLATT]["D6"] == "P"
    assert cv[CM.TAB_ERHEBUNG] == {}                      # ohne LLM keine Beurteilung


# ---- End-to-End über den Service (DB + Download) -------------------------- #

@pytest.fixture
def app(tmp_path):
    from app.shared.database import SessionLocal
    db_path = str(tmp_path / "sba.db").replace("\\", "/")

    class _Cfg(Config):
        DATABASE_URL = "sqlite:///" + db_path
        SECRET_KEY = "x"

    SessionLocal.remove()
    application = create_app(_Cfg)
    SessionLocal.remove()
    yield application
    SessionLocal.remove()


def test_end_to_end_xlsx(app):
    from app.domains.interview.models import InterviewSession
    from app.shared.database import SessionLocal
    with app.app_context():
        ps = app.projekt_service
        projekt = ps.create_projekt(org_id=1, name="BKI Test 2", verwaltungseinheit="Justiz")
        erg = ps.add_ergebnis(projekt.id, "projektinitialisierungsauftrag", created_by="Helene Digital")
        db = SessionLocal()
        db.add(InterviewSession(method_id="hermes_pia", project_name="BKI Test 2", org_id=1,
                                created_by="Helene Digital", ergebnis_id=erg.id,
                                answers_json=json.dumps({"ausgangslage": {"extracted": {"text": "x"}}})))
        db.commit()
        svc = app.schutzbedarf_service
        svc.erzeuge_entwurf(projekt)                       # keyless LLM -> deterministisch
        wb = openpyxl.load_workbook(svc.generate_xlsx(projekt))
        assert _formeln(wb[CM.TAB_ERHEBUNG]) == 127
        assert wb[CM.TAB_DECKBLATT]["D6"].value == "BKI Test 2"
